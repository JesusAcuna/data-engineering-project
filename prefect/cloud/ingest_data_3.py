# Webscraping
import requests
import os
from collections import defaultdict
from bs4 import BeautifulSoup
# Making dataframe
from read_parameters_2 import read_excel_parameters
import pandas as pd
# Loading workbook from URL
from openpyxl import load_workbook
from io import BytesIO
import urllib
# Prefect
from prefect import flow,task
from prefect_gcp.cloud_storage import GcsBucket

@task(log_prints=True)
def convert_month_to_number(value_name):
    Months=['enero','febrero','marzo','abril','mayo','junio','julio','agosto','setiembre','septiembre','octubre','noviembre','diciembre']
    Months_map=['01','02','03','04','05','06','07','08','09','09','10','11','12']

    months_dict=dict(zip(Months, Months_map))

    names_dict_new=[]

    for month in Months:
        if month in value_name.lower():
            names_dict_new.append(months_dict.get(month,"There's no value for that month in 'months_dict'"))    
    
    num_dict_new = ""
    for i in names_dict_new:
        num_dict_new=num_dict_new+i+"_"
        
    return num_dict_new[:-1]

@task(log_prints=True,retries=3)
def load_workbook_from_url(url):
    response_content = urllib.request.urlopen(url).read()

    return response_content, load_workbook(filename = BytesIO(response_content))

@task(log_prints=True)
def make_df(path,category,dep,year,max_column):
    [skiprows,names,nrows,skipfooter,usecols]=read_excel_parameters(year=year,
                                                                    category=category,
                                                                    dep=dep,
                                                                    max_column=max_column)        
    df=pd.read_excel(path,
                     sheet_name=dep,
                     skiprows=skiprows,
                     usecols=usecols,
                     names=names,
                     nrows=nrows,
                     skipfooter=skipfooter,
                     engine='openpyxl',
                     header=None)
    return df

@task(log_prints=True)
def transform_df(df):
    df.replace(['*','(SD)'], 0 , inplace=True)
    df.fillna(method="ffill",inplace=True)
    
    return df

@task(log_prints=True)
def save_excel_file(response_content,category,year,month_ranges):
    excel_file_name=f"{category}_{year}_{month_ranges}"
    path=f"./data/{year}/{excel_file_name}.xlsx"
    open(path, "wb").write(response_content)
    print(f"{path} file saved")
    
    return path

@task(log_prints=True,retries=3)
def webscraping(URL,years):
    html_text=requests.get(URL).text
    soup=BeautifulSoup(html_text,'lxml')
    block=soup.find("main",class_="l-bl block block-main-content") 
    li_tags= block.find_all('li')

    #Filters
    first_filter = "Indicadores"
    second_filter=[str(i)for i in years]

    #List of names and URLs
    names_dict=defaultdict(list)
    urls_dict=defaultdict(list)

    for li_tag in li_tags:
        li_text = li_tag.a.text.replace("\xa0", " ")
        if first_filter in li_text:
            if any((match := substring) in li_text for substring in second_filter):
                names_dict[match].append(li_text)
                li_url = li_tag.a['href']
                urls_dict[match].append(li_url)    

    return names_dict,urls_dict


@task(log_prints=True)
def write_local(df,df_file_name,year):
    path=f"./data_sheet/{year}/{df_file_name}.parquet"
    df.to_parquet(path,compression="gzip")
    return path

@task(log_prints=True)
def write_gcs(path,block):
    #from prefect_gcp.cloud_storage import GcsBucketBucket        
    block.upload_from_path(from_path=path,to_path=path)

@flow(name="Ingest Flow")
def main_flow():
    

    URL="https://web.ins.gob.pe/es/alimentacion-y-nutricion/vigilancia-alimentaria-y-nutricional/vigilancia-del-sistema-de-informacion-del-estado-nutricional-en-%20EESS" 
    years=[2022]
    
    # Block connection
    #connection_block=SqlAlchemyConnector.load("postgres-connector")
    gcs_block= GcsBucket.load("zoom-gcs")

    # Webscraping
    names_dict,urls_dict=webscraping(URL,years)

    # Make directory for xlsx files
    if not os.path.exists("./data"):
        os.mkdir("./data")  
    # Make directory for sheets of xlsx files
    if not os.path.exists("./data_sheet"):
        os.mkdir("./data_sheet")  

    years_found = names_dict.keys() #urls_dict.keys()
    
    for year in years_found:
        print(f"{year}:")

        # Make directory xlsx files
        if not os.path.exists(f"./data/{year}"):
            os.mkdir(f"./data/{year}")  

        # Make directory  for shee of xlsx files
        if not os.path.exists(f"./data_sheet/{year}"):
            os.mkdir(f"./data_sheet/{year}")  

        value_name_list=names_dict.get(year,"there's no value_name for that year")
        value_url_list=urls_dict.get(year,"there's no value_url for that year")

        for i,value_name in enumerate(value_name_list):
            
            # Convert month to number
            month_ranges = convert_month_to_number(value_name)

            if "Gestantes" in value_name:
                category="pregnant_women"
                
                if "web.ins.gob.pe" in value_url_list[i]:
                    print(f"{category}_if")
                    URL=value_url_list[i]
                    print(f"URL: {URL}")
                    
                    # Load workbook
                    response_content , wb = load_workbook_from_url(URL)
                    # Saving file
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    write_gcs(path=path,gcs_block=gcs_block)
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        # Making df 
                        df = make_df(path,category,dep,year,max_column)
                        # Transforming df
                        df = transform_df(df)                        
                        #display(df)
                        # Uploading df
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    
                        path_sheet=write_local(df,df_file_name,year)
                        write_gcs(path=path_sheet,gcs_block=gcs_block)
                    print()
                    continue
                else:
                    print(f"{category}_else")
                    URL="https://web.ins.gob.pe"+value_url_list[i]
                    print(f"URL: {URL}")
                    
                    # Load workbook
                    response_content , wb = load_workbook_from_url(URL)
                    # Saving file
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    write_gcs(path=path,block=gcs_block)
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        # Making df 
                        df = make_df(path,category,dep,year,max_column)
                        # Transforming df
                        df = transform_df(df)                        
                        #display(df)
                        # Uploading df
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    
                        path_sheet=write_local(df,df_file_name,year)
                        write_gcs(path=path_sheet,block=gcs_block)
                    print()
                    #continue
                    break

            elif ("Niños Venezolanos" in value_name) or ("Niños Extranjeros" in value_name):
                category="foreign_children"

                if "web.ins.gob.pe" in value_url_list[i]:
                    print(f"{category}_if")
                    URL=value_url_list[i]
                    print(f"URL: {URL}")
                    
                    # Load workbook
                    response_content , wb = load_workbook_from_url(URL)
                    # Saving file
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    write_gcs(path=path,block=gcs_block)
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        # Making df 
                        df = make_df(path,category,dep,year,max_column)
                        # Transforming df
                        df = transform_df(df)                        
                        #display(df)
                        # Uploading df
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    
                        path_sheet=write_local(df,df_file_name,year)
                        write_gcs(path=path_sheet,block=gcs_block)
                    print()
                    continue
                else:
                    print(f"{category}_else")
                    URL="https://web.ins.gob.pe"+value_url_list[i]
                    print(f"URL: {URL}")
                    
                    # Load workbook
                    response_content , wb = load_workbook_from_url(URL)
                    # Saving file
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    write_gcs(path=path,block=gcs_block)
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        # Making df 
                        df = make_df(path,category,dep,year,max_column)
                        # Transforming df
                        df = transform_df(df)                        
                        #display(df)
                        # Uploading df
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    
                        path_sheet=write_local(df,df_file_name,year)
                        write_gcs(path=path_sheet,block=gcs_block)
                    print()
                    continue
                    
            elif "Niños VRAEM" in value_name:
                category="VRAEM_children"

                if "web.ins.gob.pe" in value_url_list[i]:
                    print(f"{category}_if")
                    URL=value_url_list[i]
                    print(f"URL: {URL}")
                    
                    # Load workbook
                    response_content , wb = load_workbook_from_url(URL)
                    # Saving file
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    write_gcs(path=path,block=gcs_block)
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        # Making df 
                        df = make_df(path,category,dep,year,max_column)
                        # Transforming df
                        df = transform_df(df)                        
                        #display(df)
                        # Uploading df
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    
                        path_sheet=write_local(df,df_file_name,year)
                        write_gcs(path=path_sheet,block=gcs_block)
                    print()
                    continue
                else:
                    print(f"{category}_else")
                    URL="https://web.ins.gob.pe"+value_url_list[i]
                    print(f"URL: {URL}")
                    
                    # Load workbook
                    response_content , wb = load_workbook_from_url(URL)
                    # Saving file
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    write_gcs(path=path,block=gcs_block)
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        # Making df 
                        df = make_df(path,category,dep,year,max_column)
                        # Transforming df
                        df = transform_df(df)                        
                        #display(df)
                        # Uploading df
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    
                        path_sheet=write_local(df,df_file_name,year)
                        write_gcs(path=path_sheet,block=gcs_block)
                    print()
                    continue
            else:
                category="children"
                
                if "web.ins.gob.pe" in value_url_list[i]:
                    print(f"{category}_if")
                    URL=value_url_list[i]
                    print(f"URL: {URL}")
                    
                    # Load workbook
                    response_content , wb = load_workbook_from_url(URL)
                    # Saving file
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    write_gcs(path=path,block=gcs_block)
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        # Making df 
                        df = make_df(path,category,dep,year,max_column)
                        # Transforming df
                        df = transform_df(df)                        
                        #display(df)
                        # Uploading df
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    
                        path_sheet=write_local(df,df_file_name,year)
                        write_gcs(path=path_sheet,block=gcs_block)
                    print()
                    continue
                elif "gob.pe" in value_url_list[i]:
                    print(f"{category}_elif")
                    URL="https://cdn.www.gob.pe/uploads/document/file/3566498/1.Indic%20Ni%C3%B1os%20a%20Junio%202022%20-%20PERU.xlsx?v=1661961860"
                    print(f"URL: {URL}")
                    
                    # Load workbook
                    response_content , wb = load_workbook_from_url(URL)
                    # Saving file
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    write_gcs(path=path,block=gcs_block)
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        # Making df 
                        df = make_df(path,category,dep,year,max_column)
                        # Transforming df
                        df = transform_df(df)                        
                        #display(df)
                        # Uploading df
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    
                        path_sheet=write_local(df,df_file_name,year)
                        write_gcs(path=path_sheet,block=gcs_block)
                    print()
                    continue
                else:
                    print(f"{category}_else")
                    URL="https://web.ins.gob.pe"+value_url_list[i]
                    print(f"URL: {URL}")
                    
                    # Load workbook
                    response_content , wb = load_workbook_from_url(URL)
                    # Saving file
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    write_gcs(path=path,block=gcs_block)
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        # Making df 
                        df = make_df(path,category,dep,year,max_column)
                        # Transforming df
                        df = transform_df(df)                        
                        #display(df)
                        # Uploading df
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    
                        path_sheet=write_local(df,df_file_name,year)
                        write_gcs(path=path_sheet,block=gcs_block)
                    print()
                    continue
        print()
    print("End of flow")

if __name__ =='__main__':
    main_flow()
