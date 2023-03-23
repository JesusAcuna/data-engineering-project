#main
import requests
import os
from collections import defaultdict
from bs4 import BeautifulSoup
#make_df
from read_parameters_2 import read_excel_parameters
import pandas as pd
#load_workbook_from_url
from openpyxl import load_workbook
from io import BytesIO
import urllib
#
import argparse
#
from sqlalchemy import create_engine




def convert_month_to_number(value_name):
    Months=['enero','febrero','marzo','abril','mayo','junio','julio','agosto','setiembre','septiembre','octubre','noviembre','diciembre']
    Months_map=['01','02','03','04','05','06','07','08','09','09','10','11','12']

    months_dict=dict(zip(Months, Months_map))

    names_dict_new=[]

    for month in Months:
        if month in value_name.lower():
            names_dict_new.append(months_dict.get(month,"There's no value for that month in 'months_dict'"))    
    
    num_dict_new = ""
    for i in names_dict_new:
        num_dict_new=num_dict_new+i+"_"
        
    return num_dict_new[:-1]



def load_workbook_from_url(url):
    response_content = urllib.request.urlopen(url).read()
    return response_content, load_workbook(filename = BytesIO(response_content))



def make_df(path,category,dep,year,max_column):
    
    [skiprows,names,nrows,skipfooter,usecols]=read_excel_parameters(year=year,
                                                                    category=category,
                                                                    dep=dep,
                                                                    max_column=max_column)
    
    
    df=pd.read_excel(path,
                     sheet_name=dep,
                     skiprows=skiprows,
                     usecols=usecols,
                     names=names,
                     nrows=nrows,
                     skipfooter=skipfooter,
                     engine='openpyxl',
                     header=None)

    df.replace(['*','(SD)'], 0 , inplace=True)
    df.fillna(method="ffill",inplace=True)
    
    return df


def save_excel_file(response_content,category,year,month_ranges):
    excel_file_name=f"{category}_{year}_{month_ranges}"
    path=f"./data/{year}/{excel_file_name}.xlsx"
    open(path, "wb").write(response_content)
    
    print(f"\t{path} file saved")
    
    return path



#def main(URL,years):
def main(params):

    user = params.user
    password = params.password
    host = params.host
    port = params.port
    db =  params.db
    URL = params.URL
    years= params.years

    ###
    engine = create_engine(f"postgresql://{user}:{password}@{host}:{port}/{db}")
    engine.connect()

    ### WebScraping
    html_text=requests.get(URL).text
    soup=BeautifulSoup(html_text,'lxml')
    block=soup.find("main",class_="l-bl block block-main-content") 
    li_tags= block.find_all('li')

    first_filter = "Indicadores"

    second_filter=[str(i)for i in years]


    names_dict=defaultdict(list)
    urls_dict=defaultdict(list)

    for li_tag in li_tags:
        li_text = li_tag.a.text.replace("\xa0", " ")
        if first_filter in li_text:
            if any((match := substring) in li_text for substring in second_filter):
                names_dict[match].append(li_text)
                li_url = li_tag.a['href']
                urls_dict[match].append(li_url)         
    ###
            
    if not os.path.exists("./data"):
        os.mkdir("./data")  

    years = names_dict.keys() #urls_dict.keys()
    
    for year in years:
        print(f"{year}:")
        if not os.path.exists(f"./data/{year}"):
            os.mkdir(f"./data/{year}")  

        value_name_list=names_dict.get(year,"there's no value_name for that year")
        value_url_list=urls_dict.get(year,"there's no value_url for that year")

        for i,value_name in enumerate(value_name_list):
            
            month_ranges = convert_month_to_number(value_name)

            if "Gestantes" in value_name:
                category="pregnant_women"
                
                if "web.ins.gob.pe" in value_url_list[i]:
                    print(f"\t{category}_if")
                    URL=value_url_list[i]
                    print(f"\tURL: {URL}")
                    
                    response_content , wb = load_workbook_from_url(URL)
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        df = make_df(path,category,dep,year,max_column)
                        #display(df)
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    

                        df.head(n=0).to_sql(name=df_file_name,con=engine,if_exists="replace")
                        df.to_sql(name=df_file_name, con=engine, if_exists='append')
                        print(f"\t\t'{df_file_name}' table imported")

                    print()
                    continue
                else:
                    print(f"\t{category}_else")
                    URL="https://web.ins.gob.pe"+value_url_list[i]
                    print(f"\tURL: {URL}")
                    
                    response_content , wb =load_workbook_from_url(URL)
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        #print(dep,max_column)
                        df = make_df(path,category,dep,year,max_column)                        
                        #display(df)
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    

                        df.head(n=0).to_sql(name=df_file_name,con=engine,if_exists="replace")
                        df.to_sql(name=df_file_name, con=engine, if_exists='append')
                        print(f"\t\t'{df_file_name}' table imported")
                    print()
                    #continue
                    break

            elif ("Niños Venezolanos" in value_name) or ("Niños Extranjeros" in value_name):
                category="foreign_children"

                if "web.ins.gob.pe" in value_url_list[i]:
                    print(f"\t{category}_if")
                    URL=value_url_list[i]
                    print(f"\tURL: {URL}")
                    
                    response_content , wb =load_workbook_from_url(URL)
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        df = make_df(path,category,dep,year,max_column)
                        #display(df)
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    

                        df.head(n=0).to_sql(name=df_file_name,con=engine,if_exists="replace")
                        df.to_sql(name=df_file_name, con=engine, if_exists='append')                        
                        print(f"\t\t'{df_file_name}' table imported")
                    print()
                    continue
                else:
                    print(f"\t{category}_else")
                    URL="https://web.ins.gob.pe"+value_url_list[i]
                    print(f"\tURL: {URL}")
                    
                    response_content , wb =load_workbook_from_url(URL)
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        df = make_df(path,category,dep,year,max_column)
                        #display(df)
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    

                        df.head(n=0).to_sql(name=df_file_name,con=engine,if_exists="replace")
                        df.to_sql(name=df_file_name, con=engine, if_exists='append')                        
                        print(f"\t\t'{df_file_name}' table imported")
                    print()
                    continue
                    
            elif "Niños VRAEM" in value_name:
                category="VRAEM_children"

                if "web.ins.gob.pe" in value_url_list[i]:
                    print(f"\t{category}_if")
                    URL=value_url_list[i]
                    print(f"\tURL: {URL}")
                    
                    response_content , wb =load_workbook_from_url(URL)
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        df = make_df(path,category,dep,year,max_column)
                        #display(df)
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    

                        df.head(n=0).to_sql(name=df_file_name,con=engine,if_exists="replace")
                        df.to_sql(name=df_file_name, con=engine, if_exists='append')
                        print(f"\t\t'{df_file_name}' table imported")
                    print()
                    continue
                else:
                    print(f"\t{category}_else")
                    URL="https://web.ins.gob.pe"+value_url_list[i]
                    print(f"\tURL: {URL}")
                    
                    response_content , wb =load_workbook_from_url(URL)
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        df = make_df(path,category,dep,year,max_column)
                        #display(df)
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    

                        df.head(n=0).to_sql(name=df_file_name,con=engine,if_exists="replace")
                        df.to_sql(name=df_file_name, con=engine, if_exists='append')
                        print(f"\t\t'{df_file_name}' table imported")
                    print()
                    continue
            else:
                category="children"
                
                if "web.ins.gob.pe" in value_url_list[i]:
                    print(f"\t{category}_if")
                    URL=value_url_list[i]
                    print(f"\tURL: {URL}")
                    
                    response_content , wb =load_workbook_from_url(URL)
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        df = make_df(path,category,dep,year,max_column)
                        #display(df)
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    

                        df.head(n=0).to_sql(name=df_file_name,con=engine,if_exists="replace")
                        df.to_sql(name=df_file_name, con=engine, if_exists='append')
                        print(f"\t\t'{df_file_name}' table imported")
                    print()
                    continue
                elif "gob.pe" in value_url_list[i]:
                    print(f"\t{category}_elif")
                    URL="https://cdn.www.gob.pe/uploads/document/file/3566498/1.Indic%20Ni%C3%B1os%20a%20Junio%202022%20-%20PERU.xlsx?v=1661961860"
                    print(f"\tURL: {URL}")
                    
                    response_content , wb =load_workbook_from_url(URL)
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        df = make_df(path,category,dep,year,max_column)
                        #display(df)
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    

                        df.head(n=0).to_sql(name=df_file_name,con=engine,if_exists="replace")
                        df.to_sql(name=df_file_name, con=engine, if_exists='append')
                        print(f"\t\t'{df_file_name}' table imported")
                    print()
                    continue
                else:
                    print(f"\t{category}_else")
                    URL="https://web.ins.gob.pe"+value_url_list[i]
                    print(f"\tURL: {URL}")
                    
                    response_content , wb =load_workbook_from_url(URL)
                    path = save_excel_file(response_content,category,year,month_ranges)                    
                    sheets= wb.sheetnames[1:]

                    for sheet in sheets:
                        dep=sheet
                        max_column =wb[sheet].max_column
                        df = make_df(path,category,dep,year,max_column)
                        #display(df)
                        df_file_name=f"{category}_{dep}_{year}_{month_ranges}"    

                        df.head(n=0).to_sql(name=df_file_name,con=engine,if_exists="replace")
                        df.to_sql(name=df_file_name, con=engine, if_exists='append')
                        print(f"\t\t'{df_file_name}' table imported")
                    print()
                    continue

        print()
    print("End of program")


if __name__ =='__main__':

    parser = argparse.ArgumentParser(description='Ingest xlsx data to Postgres') 
    parser.add_argument('--user',help='user name for postgresql')
    parser.add_argument('--password',help='password for postgresql')
    parser.add_argument('--host',help='host for postresql')
    parser.add_argument('--port',help='port for postgresql')
    parser.add_argument('--db',help='database name for postresql')
    parser.add_argument('--URL',help='URL of the csv file')
    parser.add_argument('--years',
                        nargs="*",
                        type=int,
                        default=[],
                        help='list of ranges of years')

    args = parser.parse_args()

    main(args)
    
"""
years=list(range(2022,2023))
URL="https://web.ins.gob.pe/es/alimentacion-y-nutricion/vigilancia-alimentaria-y-nutricional/vigilancia-del-sistema-de-informacion-del-estado-nutricional-en-%20EESS"
main(URL,years=years)
"""

# Python
""" 
URL="https://web.ins.gob.pe/es/alimentacion-y-nutricion/vigilancia-alimentaria-y-nutricional/vigilancia-del-sistema-de-informacion-del-estado-nutricional-en-%20EESS"
python ingest_data.py \
  --user root \
  --password root \
  --host localhost \
  --port 5432 \
  --db ny_taxi \
  --URL ${URL} \
  --years 2022
"""

# Container
    # build
#docker build --rm -t sien_ingest:v001 .
    # entrypoing bash
# docker run --rm -it --entrypoint bash sien_ingest:v001
    # run
""" 
URL="https://web.ins.gob.pe/es/alimentacion-y-nutricion/vigilancia-alimentaria-y-nutricional/vigilancia-del-sistema-de-informacion-del-estado-nutricional-en-%20EESS"
docker run --rm -it \
   --network 2-docker-sql_default \
   sien_ingest:v001 \
   --user root \
   --password root \
   --host 2-docker-sql-pgdatabase-1 \
   --port 5432 \
   --db ny_taxi \
   --URL ${URL} \
   --years 2022
"""